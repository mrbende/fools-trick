#!/usr/bin/env python3
"""Write a run's scorecard to an .xlsx on disk -- the always-available report (no cloud needed).

Reuses report.py's aggregation so the spreadsheet mirrors the CLI scorecard exactly: one sheet
per axis, plus a Run Info sheet from the manifest. This is the on-disk fallback and default
artifact; Google Sheets (export_sheets.py) is the optional cloud upgrade when creds exist.

  bench/export_xlsx.py --dir /tmp/fools-trick/bench --stamp 20260825-110959 [--out file.xlsx]
"""
import argparse, json, os, sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import report as report_mod


def _rows(bench_dir, stamp):
    return (report_mod.speed_rows(bench_dir, stamp)
            + report_mod.capability_rows(bench_dir, stamp)
            + report_mod.jsonl_rows(bench_dir, stamp))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    ap.add_argument("--stamp", required=True)
    ap.add_argument("--out")
    a = ap.parse_args()
    rows = _rows(a.dir, a.stamp)
    if not rows:
        sys.stderr.write(f"[xlsx] no results for stamp {a.stamp}\n")
        return 0

    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    order = ["speed", "capability", "code/tools", "safety", "delegation", "long-context"]
    by_axis = {}
    for axis, node, suite, metric in rows:
        by_axis.setdefault(axis, []).append([node, suite, metric])

    first = True
    for axis in [x for x in order if x in by_axis] + [x for x in by_axis if x not in order]:
        ws = wb.active if first else wb.create_sheet()
        ws.title = axis.replace("/", "-")[:31]
        first = False
        ws.append(["node", "suite", "result"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in by_axis[axis]:
            ws.append(r)
        for col, width in (("A", 10), ("B", 40), ("C", 44)):
            ws.column_dimensions[col].width = width

    mf_path = os.path.join(a.dir, f"manifest-{a.stamp}.json")
    if os.path.exists(mf_path):
        mf = json.load(open(mf_path))
        git = mf.get("harness_git", {})
        nodes = mf.get("nodes", {})
        ws = wb.create_sheet("Run Info")
        for k, v in [("stamp", mf.get("stamp")), ("created", mf.get("created")),
                     ("size", mf.get("size")), ("seed", mf.get("seed")),
                     ("harness git sha", git.get("sha")), ("harness dirty", git.get("dirty")),
                     ("worker served", (nodes.get("worker") or {}).get("served_id")),
                     ("orchestrator served", (nodes.get("orchestrator") or {}).get("served_id"))]:
            ws.append([k, str(v)])
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 44

    out = a.out or os.path.join(a.dir, f"scorecard-{a.stamp}.xlsx")
    wb.save(out)
    print(f"xlsx -> {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
